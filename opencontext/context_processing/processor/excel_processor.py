#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Copyright (c) 2025 Beijing Volcano Engine Technology Co., Ltd.
# SPDX-License-Identifier: Apache-2.0

"""
Excel Processor with advanced cell-level chunking and metadata extraction
Handles .xlsx and .xls files with sheet-level and cell-level processing
"""

import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from opencontext.context_processing.processor.base_processor import BaseContextProcessor
from opencontext.models.context import (
    ProcessedContext,
    RawContextProperties,
    ContextProperties,
    ExtractedData,
    Chunk,
)
from opencontext.models.enums import ContextSource, ContextType, ContentFormat
from opencontext.storage.global_storage import get_storage
from opencontext.utils.logging_utils import get_logger

logger = get_logger(__name__)


class ExcelProcessor(BaseContextProcessor):
    """
    Advanced Excel processor with sheet-level and cell-level processing.
    
    Features:
    - Automatic sheet detection and processing
    - Cell-level metadata (formulas, formats, comments)
    - Table structure detection
    - Data type inference
    - Header detection
    - Summary statistics per sheet
    """

    def __init__(self):
        from opencontext.config.global_config import get_config
        config = get_config("processing.excel_processor") or {}
        super().__init__(config)
        
        # Configuration
        self._enabled = config.get("enabled", True)
        self._max_rows_per_chunk = config.get("max_rows_per_chunk", 100)
        self._extract_formulas = config.get("extract_formulas", True)
        self._extract_comments = config.get("extract_comments", True)
        self._detect_tables = config.get("detect_tables", True)
        
        logger.info("ExcelProcessor initialized")

    def get_name(self) -> str:
        return "excel_processor"

    def get_description(self) -> str:
        return "Advanced Excel processor with cell-level chunking and metadata extraction"

    @staticmethod
    def get_supported_formats() -> List[str]:
        return [".xlsx", ".xls"]

    def can_process(self, context: Any) -> bool:
        """Check if this processor can handle the context"""
        if not self._enabled:
            return False
            
        if not isinstance(context, RawContextProperties):
            return False
            
        if context.source != ContextSource.LOCAL_FILE:
            return False
            
        if not context.content_path:
            return False
            
        file_path = Path(context.content_path)
        return file_path.exists() and file_path.suffix.lower() in self.get_supported_formats()

    def process(self, context: RawContextProperties) -> List[ProcessedContext]:
        """Process Excel file and extract structured data"""
        try:
            file_path = Path(context.content_path)
            logger.info(f"Processing Excel file: {file_path}")
            
            # Read Excel file with openpyxl for full metadata access
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            processed_contexts = []
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Extract sheet-level metadata
                sheet_metadata = self._extract_sheet_metadata(sheet, file_path, sheet_name)
                
                # Convert sheet to pandas DataFrame for easier processing
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # Create chunks from the sheet
                chunks = self._create_chunks_from_sheet(sheet, df, sheet_name)
                
                # Create processed context for this sheet
                processed_context = self._create_processed_context(
                    context, chunks, sheet_metadata, sheet_name
                )
                
                if processed_context:
                    processed_contexts.append(processed_context)
            
            logger.info(f"Successfully processed {len(processed_contexts)} sheets from {file_path}")
            return processed_contexts
            
        except Exception as e:
            logger.exception(f"Error processing Excel file {context.content_path}: {e}")
            return []

    def _extract_sheet_metadata(
        self, sheet: Any, file_path: Path, sheet_name: str
    ) -> Dict[str, Any]:
        """Extract metadata from Excel sheet"""
        metadata = {
            "file_name": file_path.name,
            "file_path": str(file_path),
            "sheet_name": sheet_name,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "dimensions": sheet.dimensions,
        }
        
        # Try to detect table structures
        if self._detect_tables:
            metadata["tables"] = self._detect_table_ranges(sheet)
        
        # Extract comments if enabled
        if self._extract_comments:
            comments = []
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.comment:
                        comments.append({
                            "cell": cell.coordinate,
                            "comment": cell.comment.text
                        })
            if comments:
                metadata["comments"] = comments
        
        return metadata

    def _detect_table_ranges(self, sheet: Any) -> List[Dict[str, Any]]:
        """Detect table structures in the sheet"""
        tables = []
        
        # Look for Excel tables
        if hasattr(sheet, 'tables'):
            for table_name, table in sheet.tables.items():
                tables.append({
                    "name": table_name,
                    "range": table.ref,
                    "type": "excel_table"
                })
        
        return tables

    def _create_chunks_from_sheet(
        self, sheet: Any, df: pd.DataFrame, sheet_name: str
    ) -> List[Chunk]:
        """Create chunks from Excel sheet data"""
        chunks = []
        
        # Chunk 1: Header information
        if not df.empty:
            header_text = f"Sheet: {sheet_name}\n"
            header_text += f"Columns: {', '.join([str(col) for col in df.columns])}\n"
            header_text += f"Rows: {len(df)}, Columns: {len(df.columns)}\n"
            
            # Add data type information
            dtypes_str = "\n".join([f"{col}: {dtype}" for col, dtype in df.dtypes.items()])
            header_text += f"\nData Types:\n{dtypes_str}\n"
            
            # Add summary statistics for numeric columns
            numeric_cols = df.select_dtypes(include=['number']).columns
            if len(numeric_cols) > 0:
                header_text += f"\nNumeric Summary:\n"
                for col in numeric_cols:
                    header_text += f"{col}: min={df[col].min()}, max={df[col].max()}, mean={df[col].mean():.2f}\n"
            
            chunks.append(Chunk(
                text=header_text,
                chunk_index=0,
                keywords=[sheet_name, "header", "metadata"],
            ))
        
        # Chunk 2+: Data rows in batches
        for i in range(0, len(df), self._max_rows_per_chunk):
            chunk_df = df.iloc[i:i + self._max_rows_per_chunk]
            
            # Convert to text representation
            chunk_text = f"Sheet: {sheet_name} (Rows {i+1} to {i+len(chunk_df)})\n"
            chunk_text += chunk_df.to_string(index=False) + "\n"
            
            # Extract formulas if enabled
            if self._extract_formulas:
                formulas = self._extract_formulas_from_range(
                    sheet, i + 2, i + len(chunk_df) + 1  # +2 for 1-based index and header
                )
                if formulas:
                    chunk_text += f"\nFormulas:\n" + "\n".join(formulas)
            
            # Extract keywords from data
            keywords = [sheet_name, f"rows_{i+1}_to_{i+len(chunk_df)}"]
            
            # Add column names as keywords
            keywords.extend([str(col) for col in chunk_df.columns[:5]])  # First 5 columns
            
            chunks.append(Chunk(
                text=chunk_text,
                chunk_index=len(chunks),
                keywords=keywords,
            ))
        
        return chunks

    def _extract_formulas_from_range(
        self, sheet: Any, start_row: int, end_row: int
    ) -> List[str]:
        """Extract formulas from a range of rows"""
        formulas = []
        
        for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    col_letter = get_column_letter(col)
                    formulas.append(f"{col_letter}{row}: {cell.value}")
        
        return formulas

    def _create_processed_context(
        self,
        raw_context: RawContextProperties,
        chunks: List[Chunk],
        metadata: Dict[str, Any],
        sheet_name: str,
    ) -> Optional[ProcessedContext]:
        """Create ProcessedContext from chunks and metadata"""
        if not chunks:
            return None
        
        # Generate title and summary
        title = f"{metadata['file_name']} - {sheet_name}"
        summary = f"Excel sheet '{sheet_name}' with {metadata['max_row']} rows and {metadata['max_column']} columns"
        
        # Extract keywords
        keywords = [sheet_name, metadata['file_name'], "excel", "spreadsheet"]
        if "tables" in metadata:
            keywords.extend([table["name"] for table in metadata["tables"]])
        
        # Create ExtractedData
        extracted_data = ExtractedData(
            title=title,
            summary=summary,
            keywords=keywords,
            entities=[],
            context_type=ContextType.SEMANTIC_CONTEXT,
            confidence=90,
            importance=70,
        )
        
        # Create ContextProperties
        context_properties = ContextProperties(
            context_type=ContextType.SEMANTIC_CONTEXT,
            source=raw_context.source,
            create_time=datetime.datetime.now(),
            update_time=datetime.datetime.now(),
            content_path=raw_context.content_path,
            content_format=ContentFormat.TEXT,
            title=title,
            summary=summary,
            tags=keywords,
            additional_metadata={
                **metadata,
                "sheet_name": sheet_name,
                "processor": self.get_name(),
            },
        )
        
        # Create ProcessedContext
        processed_context = ProcessedContext(
            id=f"{raw_context.object_id}_{sheet_name}",
            properties=context_properties,
            chunks=chunks,
            extracted_data=extracted_data,
        )
        
        return processed_context
